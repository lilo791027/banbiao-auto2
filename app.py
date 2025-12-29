import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import re
from itertools import cycle

# --------------------
# 模組 1：解除合併儲存格並填入原值
# --------------------
def unmerge_and_fill(ws):
    for merged in list(ws.merged_cells.ranges):
        value = ws.cell(merged.min_row, merged.min_col).value
        ws.unmerge_cells(str(merged))
        for row in ws[merged.coord]:
            for cell in row:
                cell.value = value

# --------------------
# 模組 2：整理班表資料
# --------------------
def consolidate_selected_sheets(wb, sheet_names):
    all_data = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        unmerge_and_fill(ws)
        
        # 嘗試讀取診所名稱，若讀不到給預設值
        try:
            clinic_name = str(ws.cell(row=1, column=1).value).strip()[:4]
        except:
            clinic_name = "未知診所"

        max_row = ws.max_row
        max_col = ws.max_column
        
        for r in range(1, max_row + 1):
            for c in range(2, max_col + 1):
                cell_value = ws.cell(r, c).value
                # 簡單判斷：如果格子是日期格式，可能是日期的開頭
                if isinstance(cell_value, datetime):
                    date_val = cell_value
                    i = r + 3
                    while i <= max_row:
                        shift_type = str(ws.cell(i, c).value).strip()
                        
                        # 邊界檢查：遇到空白或非班別關鍵字停止
                        if shift_type in ["", "None"] or isinstance(ws.cell(i, c).value, datetime):
                            break
                        
                        # 抓取班別 (早/午/晚)
                        if shift_type in ["早", "午", "晚"]:
                            i += 1
                            while i <= max_row:
                                cell_v = ws.cell(i, c).value
                                if isinstance(cell_v, datetime): # 遇到下一個日期區塊
                                    break
                                
                                val = str(cell_v).strip()
                                if val in ["早", "午", "晚"]: # 遇到下一個班別
                                    break
                                
                                # 排除明顯無效的值
                                if val and val not in ["None", "nan", "="]:
                                    all_data.append([
                                        clinic_name,
                                        date_val.strftime("%Y/%m/%d"),
                                        shift_type,
                                        val
                                    ])
                                i += 1
                            i -= 1
                        i += 1
    df = pd.DataFrame(all_data, columns=["診所", "日期", "班別", "姓名"])
    return df

# --------------------
# 模組 3：建立班別分析表 (含強化的欄位對應邏輯)
# --------------------
def create_shift_analysis(df_shift: pd.DataFrame, df_emp: pd.DataFrame, shift_map: dict) -> pd.DataFrame:
    df_shift = df_shift.copy()
    df_emp = df_emp.copy()
    
    # --- 步驟 A: 強力清洗欄位名稱 (去除所有空白) ---
    # 這能解決 "職 稱" vs "職稱" 的問題
    df_shift.columns = [str(c).replace(" ", "").replace("　", "").strip() for c in df_shift.columns]
    df_emp.columns = [str(c).replace(" ", "").replace("　", "").strip() for c in df_emp.columns]
    
    # 輔助函式：模糊搜尋欄位
    def get_col_name(df, keywords):
        for col in df.columns:
            for kw in keywords:
                if kw in col:
                    return col
        return None

    # --- 步驟 B: 自動尋找對應欄位 ---
    col_map = {
        "姓名": get_col_name(df_emp, ["姓名"]),
        "編號": get_col_name(df_emp, ["編號", "工號"]),
        "職稱": get_col_name(df_emp, ["職稱", "職務", "職位"]), # 關鍵：自動找職稱相關欄位
        "部門": get_col_name(df_emp, ["部門", "單位"]),
        "分類": get_col_name(df_emp, ["分類", "類別"]),
        "特殊早班": get_col_name(df_emp, ["特殊早班", "特權"])
    }
    
    # 建立員工字典
    emp_dict = {}
    for _, row in df_emp.iterrows():
        name_col = col_map["姓名"]
        if not name_col: continue # 沒姓名欄位就跳過

        name = str(row.get(name_col, "")).strip()
        if name and name not in ["nan", "None"]:
            # 使用 .get(..., "") 防止欄位抓不到報錯
            emp_dict[name] = [
                str(row.get(col_map["編號"], "")).strip(),
                str(row.get(col_map["部門"], "")).strip(),
                str(row.get(col_map["職稱"], "")).strip(), # 這裡會抓到職稱
                str(row.get(col_map["分類"], "")).strip(),
                str(row.get(col_map["特殊早班"], "")).strip()
            ]

    # 整理班表
    shift_dict = {}
    for _, row in df_shift.iterrows():
        name = str(row.get("姓名", "")).strip()
        clinic = str(row.get("診所", "")).strip()
        date_val = row.get("日期", "")
        shift_type = str(row.get("班別", "")).strip()
        
        if not name or pd.isna(date_val): continue
        key = f"{name}|{date_val}|{clinic}"
        if key not in shift_dict: shift_dict[key] = set()
        shift_dict[key].add(shift_type)

    data_out = []
    for key, shifts in shift_dict.items():
        name, date_val, clinic = key.split("|")
        
        # 即使沒對應到員工資料，也先顯示，避免資料遺失
        emp_info = emp_dict.get(name, ["", "", "", "", ""])
        emp_id, emp_dept, emp_title, emp_category, emp_early_special = emp_info
        
        shift_parts = [s for s in ["早", "午", "晚"] if s in shifts]
        shift_type_for_code = "".join(sorted(shift_parts, key=lambda x: {"早": 1, "午": 2, "晚": 3}.get(x, 9)))
        
        class_code = get_class_code(emp_category, emp_early_special, clinic, shift_type_for_code, shift_map)
        
        data_out.append([clinic, emp_id, emp_dept, name, emp_title, date_val, shift_type_for_code, class_code])

    df_analysis = pd.DataFrame(
        data_out,
        columns=["診所", "員工編號", "所屬部門", "姓名", "職稱", "日期", "班別", "班別代碼"]
    )
    
    # 過濾無效姓名
    invalid_names = ["None", "nan", "義診", "單診", "盤點", "電打", ""]
    df_analysis = df_analysis[~df_analysis["姓名"].astype(str).str.strip().isin(invalid_names)].copy()
    
    return df_analysis

def get_class_code(emp_category, emp_early_special, clinic_name, shift_type, shift_map):
    region = "立丞" if re.search(r"立丞", str(clinic_name), re.IGNORECASE) else "板土中京"
    is_early_special = str(emp_early_special).strip().lower() in ["是", "true", "1", "checked"]

    # 1. 特殊早班
    if is_early_special and "早" in shift_type:
        if shift_type == "早": return "【員工】純早班"
        elif shift_type == "早午": return f"【員工】{region}純早、午班"
        elif shift_type == "早晚": return f"【員工】{region}純早、晚班"
        elif shift_type == "早午晚": return f"【員工】{region}純早午晚班"
    
    # 2. 一般單一早班
    if shift_type == "早":
        if "醫師" in emp_category: return "★醫師★早班"
        elif "主管" in emp_category: return "◇主管◇早班"
        elif "員工" in emp_category: return "【員工】早班"

    # 3. 早午晚 -> 全天班
    if shift_type == "早午晚":
        return f"{emp_category}{region}全天班"
    
    # 4. 其他
    base = shift_map.get(shift_type, shift_type)
    if not str(base).strip().endswith("班"): base += "班"
    return str(emp_category) + str(region) + str(base)

# --------------------
# 模組 4：建立班別總表 (含自動填補與診斷報告)
# --------------------
def create_shift_summary(df_analysis: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df_analysis.empty:
        return pd.DataFrame(), pd.DataFrame()
        
    df_analysis = df_analysis.copy()
    df_analysis["日期"] = pd.to_datetime(df_analysis["日期"], errors="coerce")
    df_analysis = df_analysis.dropna(subset=["日期"])
    all_dates = sorted(df_analysis["日期"].dt.strftime("%Y-%m-%d").unique())

    # 建立職稱對照
    emp_title_map = df_analysis[["員工編號", "姓名", "職稱"]].drop_duplicates().set_index(["員工編號", "姓名"])["職稱"].to_dict()

    # 轉置資料
    summary_dict = {}
    for _, row in df_analysis.iterrows():
        emp_id = str(row["員工編號"])
        emp_name = str(row["姓名"])
        shift_date = row["日期"].strftime("%Y-%m-%d")
        summary_dict.setdefault((emp_id, emp_name), {})[shift_date] = row["班別代碼"]

    data_out = []
    debug_list = []

    for (emp_id, emp_name), shifts in summary_dict.items():
        # 取得職稱
        raw_title = emp_title_map.get((emp_id, emp_name), "")
        title_str = str(raw_title).strip()
        
        # 排除判斷：職稱含 "醫師" 或 "兼職" 或 "PT"
        is_excluded = ("醫師" in title_str) or ("兼職" in title_str) or ("PT" in title_str.upper())
        
        # 收集診斷資訊
        debug_list.append({
            "姓名": emp_name,
            "讀到的職稱": title_str if title_str else "(空白-可能沒對應到)",
            "狀態": "❌ 不填補" if is_excluded else "✅ 自動填補",
            "原因": "是醫師/兼職" if is_excluded else "-"
        })

        leave_cycle = cycle(["{sta}", "{res}"])
        
        row = [emp_id, emp_name]
        for d in all_dates:
            val = shifts.get(d, "")
            
            # --- 強力空值判斷 ---
            # 視為空班的情況：None, nan, 空字串
            is_empty = (val is None) or (str(val).strip() in ["", "nan", "None"])
            
            if is_empty:
                if not is_excluded:
                    val = next(leave_cycle) # 填入代碼
                else:
                    val = "" # 保持空白
            
            row.append(val)
        data_out.append(row)

    cols = ["員工編號", "員工姓名"] + all_dates
    return pd.DataFrame(data_out, columns=cols), pd.DataFrame(debug_list)

# --------------------
# Streamlit 主程式
# --------------------
st.set_page_config(page_title="班表處理器(強力版)", layout="wide")
st.title("班表處理器 (強力容錯版)")
st.info("此版本會自動修正 Excel 欄位名稱差異，並提供詳細的填補診斷。")

# 檔案上傳
col1, col2 = st.columns(2)
with col1:
    shift_file = st.file_uploader("1. 上傳班表 (xlsx/xlsm)", type=["xlsx", "xlsm"])
with col2:
    employee_file = st.file_uploader("2. 上傳員工資料 (xlsx/xlsm)", type=["xlsx", "xlsm"])

if shift_file and employee_file:
    # 使用 data_only=True 讀取 Excel 計算後的值，避免讀到公式
    try:
        wb_shift = load_workbook(shift_file, data_only=True)
        wb_emp = load_workbook(employee_file, data_only=True)
    except Exception as e:
        st.error(f"檔案讀取失敗，請確認檔案未損毀: {e}")
        st.stop()

    sheets = [s for s in wb_shift.sheetnames if s not in ["彙整結果", "班別分析", "班別總表"]]
    selected_sheets = st.multiselect("選擇班表工作表", sheets)
    emp_sheet_name = st.selectbox("選擇員工資料工作表", wb_emp.sheetnames)

    if st.button("🚀 開始處理", type="primary"):
        if not selected_sheets:
            st.warning("請至少選擇一個班表！")
        else:
            with st.spinner("資料處理中..."):
                # 1. 處理班表
                df_shift = consolidate_selected_sheets(wb_shift, selected_sheets)
                
                # 2. 處理員工資料 (轉成 DataFrame)
                ws = wb_emp[emp_sheet_name]
                data = list(ws.values)
                if data:
                    cols = [str(c).strip() for c in data[0]]
                    df_emp = pd.DataFrame(data[1:], columns=cols)
                else:
                    st.error("員工資料表是空的！")
                    st.stop()

                shift_map = {"早": "早", "午": "午", "晚": "晚"}
                
                # 3. 分析
                df_analysis = create_shift_analysis(df_shift, df_emp, shift_map)
                
                # 4. 總表 (含填補)
                df_summary, df_debug = create_shift_summary(df_analysis)
            
            st.success("處理完成！")
            
            # --- 顯示診斷報告 (關鍵功能) ---
            with st.expander("🕵️‍♀️ 診斷報告：檢查誰被自動填補了？(點擊展開)", expanded=True):
                st.dataframe(df_debug, use_container_width=True)
                st.caption("說明：若【讀到的職稱】為空白，代表 Excel 欄位對應失敗；若狀態為【不填補】，代表系統判定該員為醫師或兼職。")

            st.subheader("📊 班別總表")
            st.dataframe(df_summary, use_container_width=True)

            # 下載
            with BytesIO() as output:
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df_summary.to_excel(writer, sheet_name="班別總表", index=False)
                st.download_button("📥 下載 Excel 結果", output.getvalue(), "班別總表_完整版.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
